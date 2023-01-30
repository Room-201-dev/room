from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import View, TemplateView
from .forms import WorkSpaceForm, RadioForm
from .models import WorkSpaceModel, RadioButtonModel
from django.core.mail import EmailMessage
import openpyxl



class WorkSpace(View):
    def get(self, request, *args, **kwargs):
        day = RadioButtonModel.objects.all()
        radioform = RadioForm

        return render(request, 'toyapp/work_space.html', {
            'day': day,
            'radioform': radioform
        })

    def post(self, request, *args, **kwargs):
        radioform = RadioForm(request.POST or None)

        if radioform.is_valid():
            radio = RadioButtonModel()
            user = WorkSpaceModel.bo
            radio.day = radioform.cleaned_data['day']
            radio.save()
            return redirect('top')

        return render(request, 'toyapp/work_space.html', {
            'radioform': radioform
        })


def export(request):
    day_list = RadioButtonModel.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=test.xlsx'

    ws.cell(2, 2).value = '入力内容'
    i = 3

    for days in day_list:
        ws.cell(i, 2).value = str(days.day)
        i += 1

    wb.save(response)
    return response


def top(request):
    return render(request, 'top.html')
